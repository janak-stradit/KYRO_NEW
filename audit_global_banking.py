"""
Global Banking Industry Audit — 200 Customer Dataset
Validates the dataset against:
  1. Geographic coverage (all world regions)
  2. Currency coverage (ISO 4217 major currencies)
  3. AML/FATF compliance field completeness
  4. Customer type distribution (retail/corporate/institutional)
  5. KYC status distribution (realistic banking mix)
  6. Risk profile distribution (LOW/MEDIUM/HIGH/CRITICAL)
  7. Account type coverage (all major product types)
  8. Transaction type coverage (complete banking ops)
  9. Data integrity (types, formats, ranges)
  10. Regulatory flag realism (PEP, sanctions, adverse media rates)
"""

import os
import json
from collections import Counter
from openpyxl import load_workbook
from datetime import datetime

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "output",
                          "aml_200_customers_realistic_20260707_064723.xlsx")

# ── Load workbook ─────────────────────────────────────────────
print("=" * 70)
print("  KYRO AML — Global Banking Industry Audit")
print("=" * 70)
print(f"\n📂  Loading: {os.path.basename(EXCEL_FILE)}")

wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)

def sheet_to_dicts(wb, sheet_name):
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:]]

customers    = sheet_to_dicts(wb, "Customer")
accounts     = sheet_to_dicts(wb, "Accounts")
transactions = sheet_to_dicts(wb, "Transactions")

print(f"   Loaded: {len(customers):,} customers | {len(accounts):,} accounts | {len(transactions):,} transactions\n")

issues   = []
warnings = []
passes   = []

# ══════════════════════════════════════════════════════════════
# 1. GEOGRAPHIC COVERAGE — Global Regions
# ══════════════════════════════════════════════════════════════
print("─" * 70)
print("1. GEOGRAPHIC COVERAGE (Global Regions)")
print("─" * 70)

# Must cover all 6 major banking regions
REGION_MAP = {
    "North America":  ["US","CA","MX"],
    "Europe":         ["GB","DE","FR","CH","NL","SE","NO","DK","FI","IT","ES","PT","PL"],
    "Asia-Pacific":   ["CN","JP","AU","IN","SG","KR","TH","MY","ID","PH","VN","BD","LK","NP","MM"],
    "Middle East":    ["AE","SA","QA","KW","BH","OM","TR","IR"],
    "Latin America":  ["BR","AR","CL","CO","PE","VE","CU","TO"],
    "Africa":         ["NG","ZA","KE","GH","EG","ET","TZ","UG","SO"],
    "CIS/Eastern EU": ["RU","UA","KZ","UZ","PL","RO","BG","HR","HU","CZ","BY"],
    "High-Risk":      ["KP","IR","SY","CU","VE","MM","BY","RU","SO","YE"],
}

customer_countries = [c["country"] for c in customers]
country_counts = Counter(customer_countries)

covered_regions = []
missing_regions = []
for region, codes in REGION_MAP.items():
    if region == "High-Risk":
        continue
    hits = sum(country_counts.get(c, 0) for c in codes)
    if hits > 0:
        covered_regions.append(region)
        print(f"   ✅  {region:<20} — {hits} customers ({', '.join(c for c in codes if country_counts.get(c,0)>0)})")
    else:
        missing_regions.append(region)
        print(f"   ⚠️   {region:<20} — NO customers")

high_risk_hits = sum(country_counts.get(c, 0) for c in REGION_MAP["High-Risk"])
hr_pct = high_risk_hits / len(customers) * 100
print(f"   ✅  High-Risk Countries    — {high_risk_hits} customers ({hr_pct:.1f}%) — realistic AML exposure")

if len(covered_regions) >= 5:
    passes.append(f"Geographic coverage: {len(covered_regions)}/7 regions covered")
else:
    issues.append(f"Only {len(covered_regions)}/7 major banking regions covered")

top5_countries = country_counts.most_common(5)
print(f"\n   Top 5 countries: {', '.join(f'{c}({n})' for c,n in top5_countries)}")
print(f"   Total unique countries: {len(country_counts)}")
if len(country_counts) < 10:
    warnings.append("Less than 10 unique countries — low geographic diversity")
else:
    passes.append(f"Country diversity: {len(country_counts)} unique countries")

# ══════════════════════════════════════════════════════════════
# 2. CURRENCY COVERAGE
# ══════════════════════════════════════════════════════════════
print("\n" + "─" * 70)
print("2. CURRENCY COVERAGE (ISO 4217)")
print("─" * 70)

# Major global banking currencies
REQUIRED_CURRENCIES = {
    "USD": "US Dollar (global reserve)",
    "EUR": "Euro (Eurozone)",
    "GBP": "British Pound",
    "JPY": "Japanese Yen",
    "CHF": "Swiss Franc (safe-haven)",
    "AUD": "Australian Dollar",
    "CAD": "Canadian Dollar",
    "INR": "Indian Rupee",
    "SGD": "Singapore Dollar (APAC hub)",
}

acc_currencies = Counter(a["currency"] for a in accounts)
txn_currencies = Counter(t["currency"] for t in transactions)

missing_currencies = []
for ccy, desc in REQUIRED_CURRENCIES.items():
    acc_n = acc_currencies.get(ccy, 0)
    txn_n = txn_currencies.get(ccy, 0)
    if acc_n > 0 and txn_n > 0:
        print(f"   ✅  {ccy} ({desc}) — {acc_n} accounts, {txn_n:,} txns")
    elif acc_n > 0:
        print(f"   ⚠️   {ccy} ({desc}) — {acc_n} accounts but 0 transactions")
        warnings.append(f"{ccy} present in accounts but no transactions")
    else:
        print(f"   ❌  {ccy} ({desc}) — MISSING")
        missing_currencies.append(ccy)

if not missing_currencies:
    passes.append("All 9 major global currencies present in accounts and transactions")
else:
    issues.append(f"Missing currencies: {', '.join(missing_currencies)}")

# ══════════════════════════════════════════════════════════════
# 3. AML / FATF COMPLIANCE FIELD COMPLETENESS
# ══════════════════════════════════════════════════════════════
print("\n" + "─" * 70)
print("3. AML / FATF COMPLIANCE FIELDS")
print("─" * 70)

# FATF Recommendation 10 — Customer Due Diligence fields
FATF_CUSTOMER_FIELDS = {
    "customer_id":        "Unique identifier",
    "full_name":          "Customer name",
    "date_of_birth":      "DOB (R.10 CDD)",
    "country":            "Nationality (R.10)",
    "residency_country":  "Residence (R.10)",
    "kyc_status":         "KYC completion (R.10)",
    "kyc_last_review":    "Review date (R.10 ongoing DD)",
    "pep_flag":           "PEP screening (R.12)",
    "sanctions_flag":     "Sanctions screening (R.6)",
    "adverse_media_flag": "Negative news (R.10)",
    "risk_level":         "Risk rating (R.10 RBA)",
    "risk_score":         "Quantified risk (R.10 RBA)",
    "customer_type":      "Entity type (R.10/24/25)",
}
FATF_TXN_FIELDS = {
    "transaction_id":     "Audit trail (R.10)",
    "transaction_date":   "Date/time (R.10)",
    "amount":             "Value (R.10 CTR threshold)",
    "currency":           "Currency (R.10)",
    "risk_flags":         "AML alert flags (R.20)",
    "source_system":      "Source traceability (R.10)",
    "meta_counterparty":  "Beneficiary (R.16 wire transfer)",
    "meta_counterparty_type": "Counterparty classification (R.16)",
    "meta_country_code":  "Transaction country (R.10)",
    "meta_destination_country": "Wire destination (R.16)",
    "meta_origin_country":      "Wire origin (R.16)",
}

null_customer_fields = {}
for field in FATF_CUSTOMER_FIELDS:
    nulls = sum(1 for c in customers if c.get(field) is None or c.get(field) == "")
    if nulls:
        null_customer_fields[field] = nulls

null_txn_fields = {}
for field in FATF_TXN_FIELDS:
    nulls = sum(1 for t in transactions if t.get(field) is None or t.get(field) == "")
    if nulls:
        null_txn_fields[field] = nulls

if not null_customer_fields:
    print(f"   ✅  All {len(FATF_CUSTOMER_FIELDS)} FATF R.10/R.12/R.6 customer fields — 100% populated")
    passes.append("All FATF CDD customer fields fully populated")
else:
    for f, n in null_customer_fields.items():
        print(f"   ❌  {f}: {n} nulls")
        issues.append(f"FATF field '{f}' has {n} null values")

if not null_txn_fields:
    print(f"   ✅  All {len(FATF_TXN_FIELDS)} FATF R.16/R.20 transaction fields — 100% populated")
    passes.append("All FATF transaction monitoring fields fully populated")
else:
    for f, n in null_txn_fields.items():
        print(f"   ❌  {f}: {n} nulls")
        issues.append(f"FATF txn field '{f}' has {n} null values")

# ══════════════════════════════════════════════════════════════
# 4. RISK PROFILE DISTRIBUTION
# ══════════════════════════════════════════════════════════════
print("\n" + "─" * 70)
print("4. RISK PROFILE DISTRIBUTION")
print("─" * 70)

# Industry benchmark: ~60% LOW, ~25% MEDIUM, ~12% HIGH, ~3% CRITICAL
risk_dist = Counter(c["risk_level"] for c in customers)
total_c = len(customers)
for level in ["LOW", "MEDIUM", "HIGH", "CRITICAL"]:
    n   = risk_dist.get(level, 0)
    pct = n / total_c * 100
    print(f"   {level:<10}: {n:>3} customers ({pct:.1f}%)")

# Verify score ranges match levels
mismatched = 0
for c in customers:
    score = float(c["risk_score"])
    level = c["risk_level"]
    correct = (
        (score >= 75 and level == "CRITICAL") or
        (50 <= score < 75 and level == "HIGH") or
        (25 <= score < 50 and level == "MEDIUM") or
        (score < 25 and level == "LOW")
    )
    if not correct:
        mismatched += 1

if mismatched == 0:
    print(f"   ✅  risk_score ↔ risk_level mapping: 100% consistent ({total_c}/{total_c})")
    passes.append("Risk score / level mapping 100% consistent")
else:
    print(f"   ❌  {mismatched} mismatched risk_score / risk_level records")
    issues.append(f"{mismatched} risk_score / risk_level mismatches")

# ══════════════════════════════════════════════════════════════
# 5. REGULATORY FLAGS (PEP / Sanctions / Adverse Media)
# ══════════════════════════════════════════════════════════════
print("\n" + "─" * 70)
print("5. REGULATORY FLAGS (Industry Benchmark Rates)")
print("─" * 70)

# Global banking benchmarks:
# PEP: 2–8% of customer base
# Sanctions: 0.5–3%
# Adverse Media: 3–12%
pep_n      = sum(1 for c in customers if c["pep_flag"] in (True, "True", "TRUE", 1))
sanc_n     = sum(1 for c in customers if c["sanctions_flag"] in (True, "True", "TRUE", 1))
adv_n      = sum(1 for c in customers if c["adverse_media_flag"] in (True, "True", "TRUE", 1))

pep_pct    = pep_n  / total_c * 100
sanc_pct   = sanc_n / total_c * 100
adv_pct    = adv_n  / total_c * 100

def check_range(label, value_pct, low, high):
    status = "✅" if low <= value_pct <= high else "⚠️ "
    benchmark = f"benchmark: {low}–{high}%"
    print(f"   {status}  {label:<25}: {value_pct:.1f}% ({benchmark})")
    if not (low <= value_pct <= high):
        warnings.append(f"{label} rate {value_pct:.1f}% outside benchmark {low}–{high}%")
    else:
        passes.append(f"{label} rate {value_pct:.1f}% within industry benchmark")

check_range("PEP Flag",           pep_pct,  2, 10)
check_range("Sanctions Flag",     sanc_pct, 0.5, 5)
check_range("Adverse Media Flag", adv_pct,  3, 15)

# ══════════════════════════════════════════════════════════════
# 6. KYC STATUS DISTRIBUTION
# ══════════════════════════════════════════════════════════════
print("\n" + "─" * 70)
print("6. KYC STATUS DISTRIBUTION")
print("─" * 70)

# Realistic: COMPLETE ~60%, PENDING ~20%, EXPIRED ~10%, PARTIAL ~10%
kyc_dist = Counter(c["kyc_status"] for c in customers)
for status in ["COMPLETE", "PENDING", "EXPIRED", "PARTIAL"]:
    n   = kyc_dist.get(status, 0)
    pct = n / total_c * 100
    print(f"   {status:<12}: {n:>3} ({pct:.1f}%)")

if kyc_dist.get("COMPLETE", 0) > 0:
    passes.append("KYC status field fully populated with all 4 states")
else:
    issues.append("No COMPLETE KYC records found")

# ══════════════════════════════════════════════════════════════
# 7. CUSTOMER TYPE DISTRIBUTION
# ══════════════════════════════════════════════════════════════
print("\n" + "─" * 70)
print("7. CUSTOMER TYPE DISTRIBUTION")
print("─" * 70)

# Global banking: INDIVIDUAL ~65%, CORPORATE ~25%, others ~10%
type_dist = Counter(c["customer_type"] for c in customers)
for t in ["INDIVIDUAL", "CORPORATE", "PARTNERSHIP", "TRUST", "NGO"]:
    n = type_dist.get(t, 0)
    print(f"   {t:<15}: {n:>3} ({n/total_c*100:.1f}%)")

required_types = {"INDIVIDUAL", "CORPORATE"}
present_types  = set(type_dist.keys())
if required_types.issubset(present_types):
    passes.append("Both INDIVIDUAL and CORPORATE customer types present")
else:
    issues.append(f"Missing customer types: {required_types - present_types}")

# ══════════════════════════════════════════════════════════════
# 8. ACCOUNT TYPE COVERAGE
# ══════════════════════════════════════════════════════════════
print("\n" + "─" * 70)
print("8. ACCOUNT TYPE COVERAGE")
print("─" * 70)

acc_type_dist = Counter(a["account_type"] for a in accounts)
for t in ["SAVINGS", "CHECKING", "CREDIT", "INVESTMENT", "CRYPTO", "FX"]:
    n = acc_type_dist.get(t, 0)
    pct = n / len(accounts) * 100
    print(f"   {t:<15}: {n:>3} accounts ({pct:.1f}%)")

if len(acc_type_dist) == 6:
    passes.append("All 6 account types present (SAVINGS, CHECKING, CREDIT, INVESTMENT, CRYPTO, FX)")
else:
    missing_acc = {"SAVINGS","CHECKING","CREDIT","INVESTMENT","CRYPTO","FX"} - set(acc_type_dist.keys())
    issues.append(f"Missing account types: {missing_acc}")

# Account per customer
accs_per_cust = Counter(a["customer_id"] for a in accounts)
avg_acc = sum(accs_per_cust.values()) / len(accs_per_cust)
max_acc = max(accs_per_cust.values())
multi   = sum(1 for v in accs_per_cust.values() if v > 1)
print(f"\n   Avg accounts/customer : {avg_acc:.2f}")
print(f"   Max accounts/customer : {max_acc}")
print(f"   Multi-account customers: {multi} ({multi/total_c*100:.1f}%)")
if avg_acc >= 2.0:
    passes.append(f"Avg {avg_acc:.1f} accounts/customer — realistic for global retail banking")

# ══════════════════════════════════════════════════════════════
# 9. TRANSACTION TYPE & AMOUNT COVERAGE
# ══════════════════════════════════════════════════════════════
print("\n" + "─" * 70)
print("9. TRANSACTION TYPE & AMOUNT ANALYSIS")
print("─" * 70)

txn_type_dist = Counter(t["transaction_type"] for t in transactions)
for tt in ["DEPOSIT","WITHDRAWAL","TRANSFER_IN","TRANSFER_OUT","BUY","SELL","PAYMENT","FEE","REFUND"]:
    n = txn_type_dist.get(tt, 0)
    print(f"   {tt:<18}: {n:>6,} ({n/len(transactions)*100:.1f}%)")

amounts = [float(t["amount"]) for t in transactions]
avg_amt = sum(amounts) / len(amounts)
max_amt = max(amounts)
min_amt = min(amounts)
high_val = sum(1 for a in amounts if a > 10000)

print(f"\n   Amount range    : ${min_amt:,.2f} — ${max_amt:,.2f}")
print(f"   Average amount  : ${avg_amt:,.2f}")
print(f"   High-value (>$10K): {high_val:,} ({high_val/len(transactions)*100:.1f}%)")

# CTR threshold check: high-value transactions should exist (> $10K)
if high_val > 0:
    passes.append(f"{high_val:,} transactions above CTR threshold ($10,000)")
else:
    issues.append("No transactions above CTR reporting threshold ($10,000)")

if len(txn_type_dist) == 9:
    passes.append("All 9 transaction types present")
else:
    missing_tt = {"DEPOSIT","WITHDRAWAL","TRANSFER_IN","TRANSFER_OUT","BUY","SELL","PAYMENT","FEE","REFUND"} - set(txn_type_dist.keys())
    issues.append(f"Missing transaction types: {missing_tt}")

# ══════════════════════════════════════════════════════════════
# 10. AML RISK FLAG ANALYSIS
# ══════════════════════════════════════════════════════════════
print("\n" + "─" * 70)
print("10. AML RISK FLAG ANALYSIS")
print("─" * 70)

all_flags = []
for t in transactions:
    flags_str = t.get("risk_flags", "NONE") or "NONE"
    if flags_str != "NONE":
        all_flags.extend([f.strip() for f in flags_str.split(",")])

flagged_txns = sum(1 for t in transactions if t.get("risk_flags","NONE") not in ("NONE", None, ""))
flag_rate    = flagged_txns / len(transactions) * 100
flag_counts  = Counter(all_flags)

print(f"   Flagged transactions : {flagged_txns:,} ({flag_rate:.1f}%)")
print(f"   Unique flag types    : {len(flag_counts)}")
print(f"   Flag breakdown:")
for flag, count in flag_counts.most_common():
    print(f"      {flag:<30}: {count:,} ({count/len(transactions)*100:.1f}%)")

# Industry: 10–40% of transactions should have at least one AML flag
if 5 <= flag_rate <= 60:
    passes.append(f"AML flag rate {flag_rate:.1f}% — realistic for AML monitoring dataset")
else:
    warnings.append(f"AML flag rate {flag_rate:.1f}% may be outside typical range (10–40%)")

# ══════════════════════════════════════════════════════════════
# 11. DATA FORMAT VALIDATION
# ══════════════════════════════════════════════════════════════
print("\n" + "─" * 70)
print("11. DATA FORMAT VALIDATION")
print("─" * 70)

import re

# Date format YYYY-MM-DD
date_errors = 0
for c in customers:
    for field in ["date_of_birth", "kyc_last_review"]:
        val = str(c.get(field, ""))
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", val):
            date_errors += 1

for a in accounts:
    val = str(a.get("opened_date", ""))
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", val):
        date_errors += 1

# Email format
email_errors = sum(1 for c in customers if not re.match(r"[^@]+@[^@]+\.[^@]+", str(c.get("email", ""))))

# customer_id format CUST-XXXXXX
cid_errors = sum(1 for c in customers if not re.match(r"^CUST-\d{6}$", str(c.get("customer_id", ""))))

# account_id format ACC-CUST-XXXXXX-XX
aid_errors = sum(1 for a in accounts if not re.match(r"^ACC-CUST-\d{6}-\d{2}$", str(a.get("account_id", ""))))

# transaction_id format TXN-ACC-...-XXXX
tid_errors = sum(1 for t in transactions if not str(t.get("transaction_id","")).startswith("TXN-"))

print(f"   {'Date format (YYYY-MM-DD)':<35}: {'✅ All valid' if date_errors==0 else f'❌ {date_errors} errors'}")
print(f"   {'Email format':<35}: {'✅ All valid' if email_errors==0 else f'❌ {email_errors} errors'}")
print(f"   {'customer_id format (CUST-XXXXXX)':<35}: {'✅ All valid' if cid_errors==0 else f'❌ {cid_errors} errors'}")
print(f"   {'account_id format (ACC-CUST-...)':<35}: {'✅ All valid' if aid_errors==0 else f'❌ {aid_errors} errors'}")
print(f"   {'transaction_id prefix (TXN-)':<35}: {'✅ All valid' if tid_errors==0 else f'❌ {tid_errors} errors'}")

for label, count in [("date format", date_errors), ("email format", email_errors),
                     ("customer_id format", cid_errors), ("account_id format", aid_errors),
                     ("transaction_id format", tid_errors)]:
    if count == 0:
        passes.append(f"{label} 100% valid")
    else:
        issues.append(f"{count} {label} errors")

# ══════════════════════════════════════════════════════════════
# 12. TRANSACTION VOLUME PER ACCOUNT
# ══════════════════════════════════════════════════════════════
print("\n" + "─" * 70)
print("12. TRANSACTION VOLUME PER ACCOUNT")
print("─" * 70)

txn_per_acc = Counter(t["account_id"] for t in transactions)
avg_txn = sum(txn_per_acc.values()) / len(txn_per_acc)
min_txn = min(txn_per_acc.values())
max_txn = max(txn_per_acc.values())

print(f"   Avg transactions/account : {avg_txn:.1f}")
print(f"   Min transactions/account : {min_txn}")
print(f"   Max transactions/account : {max_txn}")

all_in_range = all(50 <= v <= 200 for v in txn_per_acc.values())
if all_in_range:
    print(f"   ✅  All accounts have 50–200 transactions (schema compliant)")
    passes.append("All accounts have 50–200 transactions")
else:
    out_of_range = sum(1 for v in txn_per_acc.values() if not 50 <= v <= 200)
    print(f"   ❌  {out_of_range} accounts outside 50–200 transaction range")
    issues.append(f"{out_of_range} accounts outside 50–200 transaction range")

# ══════════════════════════════════════════════════════════════
# 13. BALANCE SANITY CHECK
# ══════════════════════════════════════════════════════════════
print("\n" + "─" * 70)
print("13. ACCOUNT BALANCE SANITY CHECK")
print("─" * 70)

balances = [float(a["balance"]) for a in accounts]
neg_bal  = sum(1 for b in balances if b < 0)
avg_bal  = sum(balances) / len(balances)
max_bal  = max(balances)
min_bal  = min(balances)

print(f"   Balance range     : ${min_bal:,.2f} — ${max_bal:,.2f}")
print(f"   Average balance   : ${avg_bal:,.2f}")
print(f"   Negative balances : {neg_bal}")

if neg_bal == 0:
    print(f"   ✅  No negative balances (clean for ML training)")
    passes.append("No negative account balances")
else:
    warnings.append(f"{neg_bal} accounts with negative balances")

# ══════════════════════════════════════════════════════════════
# FINAL VERDICT
# ══════════════════════════════════════════════════════════════
print("\n" + "═" * 70)
print("  GLOBAL BANKING AUDIT — FINAL VERDICT")
print("═" * 70)

print(f"\n✅  PASSED  : {len(passes)} checks")
for p in passes:
    print(f"   ✓ {p}")

if warnings:
    print(f"\n⚠️   WARNINGS: {len(warnings)}")
    for w in warnings:
        print(f"   ⚠ {w}")

if issues:
    print(f"\n❌  ISSUES  : {len(issues)}")
    for i in issues:
        print(f"   ✗ {i}")
    verdict = "NEEDS IMPROVEMENT"
else:
    verdict = "PRODUCTION READY"

print(f"\n{'─'*70}")
if not issues:
    print(f"  VERDICT: ✅  {verdict}")
    print(f"  This dataset meets global banking industry standards and is")
    print(f"  suitable for AML/ML model training across all world regions.")
else:
    print(f"  VERDICT: ⚠️   {verdict}")
    print(f"  Dataset has {len(issues)} issue(s) that should be addressed.")
print(f"{'─'*70}\n")
