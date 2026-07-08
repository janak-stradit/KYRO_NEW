"""
Generates a 200-customer AML dataset and saves it to Excel.
Data quality guarantees enforced before write:
  - no null / empty cells
  - no duplicate IDs across any sheet
  - referential integrity on all foreign keys
  - risk_score to risk_level mapping consistent

Usage:
    python3 generate_200_clean.py
"""

import os
import re
import sys
import uuid
import random
import time
from collections import Counter
from datetime import datetime, timedelta, timezone

from faker import Faker
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

fake = Faker()
random.seed(42)

NUM_CUSTOMERS = 200
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

COUNTRIES = [
    "US", "GB", "DE", "FR", "IN", "CN", "JP", "AU", "CA", "BR",
    "TO", "KP", "RU", "IR", "PK", "NG", "MX", "ZA", "AE", "SG",
    "CH", "NL", "SE", "NO", "DK", "FI", "IT", "ES", "PT", "PL",
]
HIGH_RISK = {"KP", "IR", "SY", "CU", "VE", "MM", "BY", "RU", "SO", "YE"}
CURRENCIES = ["USD", "EUR", "GBP", "CHF", "AUD", "CAD", "JPY", "INR", "SGD"]
KYC_STATUSES = ["COMPLETE", "PENDING", "EXPIRED", "PARTIAL"]
CUSTOMER_TYPES = ["INDIVIDUAL", "CORPORATE", "PARTNERSHIP", "TRUST", "NGO"]
ACCOUNT_TYPES = ["SAVINGS", "CHECKING", "CREDIT", "INVESTMENT", "CRYPTO", "FX"]
ACCOUNT_STATUSES = ["ACTIVE", "CLOSED", "FROZEN", "SUSPENDED"]
TXN_TYPES = [
    "DEPOSIT", "WITHDRAWAL", "TRANSFER_IN", "TRANSFER_OUT",
    "BUY", "SELL", "PAYMENT", "FEE", "REFUND",
]
SOURCE_SYSTEMS = ["bank_mcp", "card_mcp", "crypto_mcp", "swift_mcp", "internal"]
CP_TYPES = ["INDIVIDUAL", "BANK", "CORPORATE", "EXCHANGE", "UNKNOWN"]

CUSTOMER_COLS = [
    "customer_id", "full_name", "email", "phone", "date_of_birth",
    "country", "residency_country", "kyc_status", "kyc_last_review",
    "pep_flag", "sanctions_flag", "adverse_media_flag", "risk_level",
    "risk_score", "customer_type", "customer_metadata",
]
ACCOUNT_COLS = [
    "account_id", "customer_id", "account_type", "account_status",
    "currency", "balance", "opened_date", "account_metadata",
]
TXN_COLS = [
    "transaction_id", "customer_id", "account_id", "transaction_date",
    "transaction_type", "amount", "currency", "risk_flags",
    "source_system", "meta_counterparty", "meta_counterparty_type",
    "meta_location", "meta_country", "meta_country_code",
    "meta_destination_country", "meta_origin_country", "meta_source",
]
DATA_DICT_ROWS = [
    ("Customer", "customer_id", "String", "object"),
    ("Customer", "full_name", "String", "object"),
    ("Customer", "email", "String", "object"),
    ("Customer", "phone", "String", "object"),
    ("Customer", "date_of_birth", "String", "object"),
    ("Customer", "country", "String", "object"),
    ("Customer", "residency_country", "String", "object"),
    ("Customer", "kyc_status", "String", "object"),
    ("Customer", "kyc_last_review", "String", "object"),
    ("Customer", "pep_flag", "Boolean", "bool"),
    ("Customer", "sanctions_flag", "Boolean", "bool"),
    ("Customer", "adverse_media_flag", "Boolean", "bool"),
    ("Customer", "risk_level", "String", "object"),
    ("Customer", "risk_score", "Float", "float64"),
    ("Customer", "customer_type", "String", "object"),
    ("Customer", "customer_metadata", "String", "object"),
    ("Accounts", "account_id", "String", "object"),
    ("Accounts", "customer_id", "String", "object"),
    ("Accounts", "account_type", "String", "object"),
    ("Accounts", "account_status", "String", "object"),
    ("Accounts", "currency", "String", "object"),
    ("Accounts", "balance", "Float", "float64"),
    ("Accounts", "opened_date", "String", "object"),
    ("Accounts", "account_metadata", "String", "object"),
    ("Transactions", "transaction_id", "String", "object"),
    ("Transactions", "customer_id", "String", "object"),
    ("Transactions", "account_id", "String", "object"),
    ("Transactions", "transaction_date", "String", "object"),
    ("Transactions", "transaction_type", "String", "object"),
    ("Transactions", "amount", "Float", "float64"),
    ("Transactions", "currency", "String", "object"),
    ("Transactions", "risk_flags", "String", "object"),
    ("Transactions", "source_system", "String", "object"),
    ("Transactions", "meta_counterparty", "String", "object"),
    ("Transactions", "meta_counterparty_type", "String", "object"),
    ("Transactions", "meta_location", "String", "object"),
    ("Transactions", "meta_country", "String", "object"),
    ("Transactions", "meta_country_code", "String", "object"),
    ("Transactions", "meta_destination_country", "String", "object"),
    ("Transactions", "meta_origin_country", "String", "object"),
    ("Transactions", "meta_source", "String", "object"),
]


def rand_date(start_yr, end_yr):
    start = datetime(start_yr, 1, 1)
    span = (datetime(end_yr, 12, 31) - start).days
    return (start + timedelta(days=random.randint(0, span))).strftime("%Y-%m-%d")


def risk_flags(amount, country, txn_type):
    flags = []
    if amount > 10000:
        flags.append("HIGH_VALUE")
    if country in HIGH_RISK:
        flags.append("HIGH_RISK_COUNTRY")
    if txn_type in ("TRANSFER_IN", "TRANSFER_OUT") and random.random() < 0.25:
        flags.append("CROSS_BORDER")
    if random.randint(0, 23) in range(0, 6):
        flags.append("UNUSUAL_TIME")
    if random.random() < 0.05:
        flags.append("SUSPICIOUS_COUNTERPARTY")
    if random.random() < 0.03:
        flags.append("UNUSUAL_PATTERN")
    if random.random() < 0.02:
        flags.append("STRUCTURING")
    if random.random() < 0.02:
        flags.append("RAPID_MOVEMENT")
    return ", ".join(flags) if flags else "NONE"


def txn_amount():
    r = random.random()
    if r < 0.55:
        return round(random.uniform(5.0, 500.0), 2)
    elif r < 0.80:
        return round(random.uniform(500.0, 5000.0), 2)
    elif r < 0.93:
        return round(random.uniform(5000.0, 50000.0), 2)
    else:
        return round(random.uniform(50000.0, 250000.0), 2)


def make_customer(index):
    cid = f"CUST-{str(index).zfill(6)}"
    score = round(random.uniform(0, 100), 4)
    level = "CRITICAL" if score >= 75 else "HIGH" if score >= 50 else "MEDIUM" if score >= 25 else "LOW"
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    return {
        "customer_id": cid,
        "full_name": fake.name(),
        "email": fake.unique.email(),
        "phone": fake.phone_number(),
        "date_of_birth": fake.date_of_birth(minimum_age=18, maximum_age=80).strftime("%Y-%m-%d"),
        "country": random.choice(COUNTRIES),
        "residency_country": random.choice(COUNTRIES),
        "kyc_status": random.choice(KYC_STATUSES),
        "kyc_last_review": rand_date(2020, 2025),
        "pep_flag": random.random() < 0.05,
        "sanctions_flag": random.random() < 0.02,
        "adverse_media_flag": random.random() < 0.08,
        "risk_level": level,
        "risk_score": score,
        "customer_type": random.choice(CUSTOMER_TYPES),
        "customer_metadata": f"{{'entity': 'customer', 'source': 'mock_data_service', 'generated_at': '{ts}'}}",
    }


def make_accounts(customer_id):
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    result = []
    for i in range(1, random.randint(1, 5) + 1):
        result.append({
            "account_id": f"ACC-{customer_id}-{str(i).zfill(2)}",
            "customer_id": customer_id,
            "account_type": random.choice(ACCOUNT_TYPES),
            "account_status": random.choice(ACCOUNT_STATUSES),
            "currency": random.choice(CURRENCIES),
            "balance": round(random.uniform(100.0, 500000.0), 2),
            "opened_date": rand_date(2010, 2024),
            "account_metadata": f"{{'entity': 'account', 'source': 'mock_data_service', 'generated_at': '{ts}'}}",
        })
    return result


def make_transactions(customer_id, account_id):
    seen = set()
    result = []
    for _ in range(random.randint(50, 200)):
        txn_type = random.choice(TXN_TYPES)
        amount = txn_amount()
        country = random.choice(COUNTRIES)
        cp_type = random.choice(CP_TYPES)

        if cp_type in ("BANK", "CORPORATE", "EXCHANGE"):
            cp_name = fake.company()
        elif cp_type == "INDIVIDUAL":
            cp_name = fake.name()
        else:
            cp_name = "UNKNOWN ENTITY"

        dest = random.choice(COUNTRIES) if txn_type in ("TRANSFER_IN", "TRANSFER_OUT") else "N/A"
        origin = random.choice(COUNTRIES) if txn_type == "TRANSFER_IN" else "N/A"

        tid = f"TXN-{account_id}-{uuid.uuid4().hex[:12].upper()}"
        while tid in seen:
            tid = f"TXN-{account_id}-{uuid.uuid4().hex[:12].upper()}"
        seen.add(tid)

        result.append({
            "transaction_id": tid,
            "customer_id": customer_id,
            "account_id": account_id,
            "transaction_date": rand_date(2020, 2025),
            "transaction_type": txn_type,
            "amount": amount,
            "currency": random.choice(CURRENCIES),
            "risk_flags": risk_flags(amount, country, txn_type),
            "source_system": random.choice(SOURCE_SYSTEMS),
            "meta_counterparty": cp_name,
            "meta_counterparty_type": cp_type,
            "meta_location": fake.city(),
            "meta_country": fake.country(),
            "meta_country_code": country,
            "meta_destination_country": dest,
            "meta_origin_country": origin,
            "meta_source": "mock_data_service",
        })
    return result


def build_dataset(n):
    customers, accounts, transactions = [], [], []
    for i in range(1, n + 1):
        c = make_customer(i)
        customers.append(c)
        accs = make_accounts(c["customer_id"])
        accounts.extend(accs)
        for acc in accs:
            transactions.extend(make_transactions(c["customer_id"], acc["account_id"]))
    return customers, accounts, transactions


def validate(customers, accounts, transactions):
    errors = []

    cids = [c["customer_id"] for c in customers]
    aids = [a["account_id"] for a in accounts]
    tids = [t["transaction_id"] for t in transactions]

    if len(cids) != len(set(cids)):
        errors.append(f"{len(cids) - len(set(cids))} duplicate customer_ids")
    if len(aids) != len(set(aids)):
        errors.append(f"{len(aids) - len(set(aids))} duplicate account_ids")
    if len(tids) != len(set(tids)):
        errors.append(f"{len(tids) - len(set(tids))} duplicate transaction_ids")

    for i, c in enumerate(customers):
        for k, v in c.items():
            if v is None or v == "":
                errors.append(f"Customer[{i}].{k} is null/empty")

    for i, a in enumerate(accounts):
        for k, v in a.items():
            if v is None or v == "":
                errors.append(f"Account[{i}].{k} is null/empty")

    null_txn = sum(1 for t in transactions for v in t.values() if v is None or v == "")
    if null_txn:
        errors.append(f"{null_txn} null/empty fields in transactions")

    valid_cids = set(cids)
    valid_aids = set(aids)
    bad = sum(1 for a in accounts if a["customer_id"] not in valid_cids)
    if bad:
        errors.append(f"{bad} accounts with invalid customer_id")
    bad = sum(1 for t in transactions if t["customer_id"] not in valid_cids)
    if bad:
        errors.append(f"{bad} transactions with invalid customer_id")
    bad = sum(1 for t in transactions if t["account_id"] not in valid_aids)
    if bad:
        errors.append(f"{bad} transactions with invalid account_id")

    bad_risk = sum(1 for c in customers if not (
        (c["risk_score"] >= 75 and c["risk_level"] == "CRITICAL") or
        (50 <= c["risk_score"] < 75 and c["risk_level"] == "HIGH") or
        (25 <= c["risk_score"] < 50 and c["risk_level"] == "MEDIUM") or
        (c["risk_score"] < 25 and c["risk_level"] == "LOW")
    ))
    if bad_risk:
        errors.append(f"{bad_risk} customers with mismatched risk_score/risk_level")

    return errors


# Excel helpers
_HDR_FILL = PatternFill("solid", fgColor="1F3864")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN = Side(style="thin", color="B0B0B0")
_HDR_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _header(ws, n):
    for col in range(1, n + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = _HDR_ALIGN
        cell.border = _HDR_BORDER


def _autofit(ws, cap=50):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        best = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(best + 4, cap)


def write_sheet(wb, name, cols, rows):
    ws = wb.create_sheet(name)
    ws.append(cols)
    _header(ws, len(cols))
    for row in rows:
        ws.append([row.get(c) for c in cols])
    _autofit(ws)
    ws.freeze_panes = "A2"


# -- main --

print(f"Generating {NUM_CUSTOMERS} customers ...")
t0 = time.time()
customers, accounts, transactions = build_dataset(NUM_CUSTOMERS)
print(f"Done in {time.time() - t0:.1f}s — "
      f"{len(customers)} customers / {len(accounts)} accounts / {len(transactions):,} transactions")

print("Validating ...")
errors = validate(customers, accounts, transactions)
if errors:
    print("FAILED:")
    for e in errors:
        print(f"  - {e}")
    sys.exit(1)

multi_acc = sum(
    1 for cid in set(a["customer_id"] for a in accounts)
    if sum(1 for a in accounts if a["customer_id"] == cid) > 1
)
print(f"  no duplicates, no nulls, FK integrity OK")
print(f"  multi-account customers: {multi_acc}/{NUM_CUSTOMERS} ({multi_acc/NUM_CUSTOMERS*100:.0f}%)")
print("All checks passed")

ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
fname = f"aml_200_customers_realistic_{ts}.xlsx"
fpath = os.path.join(OUTPUT_DIR, fname)

print(f"Writing {fname} ...")
t1 = time.time()
wb = Workbook()
if wb.active is not None:
    wb.remove(wb.active)

write_sheet(wb, "Customer", CUSTOMER_COLS, customers)
write_sheet(wb, "Accounts", ACCOUNT_COLS, accounts)
write_sheet(wb, "Transactions", TXN_COLS, transactions)

ws_dd = wb.create_sheet("Data_Dictionary")
ws_dd.append(["Table Name", "Column Name", "Required Data Type", "Pandas Dtype"])
_header(ws_dd, 4)
for row in DATA_DICT_ROWS:
    ws_dd.append(list(row))
_autofit(ws_dd, cap=40)
ws_dd.freeze_panes = "A2"

wb.save(fpath)
size_kb = os.path.getsize(fpath) / 1024
print(f"Saved: {fpath}  ({size_kb:.0f} KB, {time.time() - t1:.1f}s)")
