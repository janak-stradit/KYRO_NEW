import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Column order matches the reference dataset (CUST-ML-TEST-001_ml_dataset.xlsx)
CUSTOMER_COLS = [
    "customer_id", "full_name", "email", "phone", "date_of_birth",
    "country", "residency_country", "kyc_status", "kyc_last_review",
    "pep_flag", "sanctions_flag", "adverse_media_flag", "risk_level",
    "risk_score", "customer_type", "customer_metadata",
]

ACCOUNT_COLS = [
    "account_id", "customer_id", "account_type", "account_status",
    "currency", "balance", "opened_date", "account_metadata",
]

TRANSACTION_COLS = [
    "transaction_id", "customer_id", "account_id", "transaction_date",
    "transaction_type", "amount", "currency", "risk_flags",
    "source_system", "meta_counterparty", "meta_counterparty_type",
    "meta_location", "meta_country", "meta_country_code",
    "meta_destination_country", "meta_origin_country", "meta_source",
]

# Shown in Data_Dictionary tab: (sheet, column, type label, pandas dtype)
DATA_DICT = [
    ("Customer", "customer_id", "String", "object"),
    ("Customer", "full_name", "String", "object"),
    ("Customer", "email", "String", "object"),
    ("Customer", "phone", "String", "object"),
    ("Customer", "date_of_birth", "String", "object"),
    ("Customer", "country", "String", "object"),
    ("Customer", "residency_country", "String", "object"),
    ("Customer", "kyc_status", "String", "object"),
    ("Customer", "kyc_last_review", "String", "object"),
    ("Customer", "pep_flag", "Boolean", "bool"),
    ("Customer", "sanctions_flag", "Boolean", "bool"),
    ("Customer", "adverse_media_flag", "Boolean", "bool"),
    ("Customer", "risk_level", "String", "object"),
    ("Customer", "risk_score", "Float", "float64"),
    ("Customer", "customer_type", "String", "object"),
    ("Customer", "customer_metadata", "String", "object"),
    ("Accounts", "account_id", "String", "object"),
    ("Accounts", "customer_id", "String", "object"),
    ("Accounts", "account_type", "String", "object"),
    ("Accounts", "account_status", "String", "object"),
    ("Accounts", "currency", "String", "object"),
    ("Accounts", "balance", "Float", "float64"),
    ("Accounts", "opened_date", "String", "object"),
    ("Accounts", "account_metadata", "String", "object"),
    ("Transactions", "transaction_id", "String", "object"),
    ("Transactions", "customer_id", "String", "object"),
    ("Transactions", "account_id", "String", "object"),
    ("Transactions", "transaction_date", "String", "object"),
    ("Transactions", "transaction_type", "String", "object"),
    ("Transactions", "amount", "Float", "float64"),
    ("Transactions", "currency", "String", "object"),
    ("Transactions", "risk_flags", "String", "object"),
    ("Transactions", "source_system", "String", "object"),
    ("Transactions", "meta_counterparty", "String", "object"),
    ("Transactions", "meta_counterparty_type", "String", "object"),
    ("Transactions", "meta_location", "String", "object"),
    ("Transactions", "meta_country", "String", "object"),
    ("Transactions", "meta_country_code", "String", "object"),
    ("Transactions", "meta_destination_country", "String", "object"),
    ("Transactions", "meta_origin_country", "String", "object"),
    ("Transactions", "meta_source", "String", "object"),
]

_HDR_FILL = PatternFill("solid", fgColor="1F3864")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_HDR_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def _apply_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = _HDR_ALIGN
        cell.border = _HDR_BORDER


def _fit_columns(ws, cap=60):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        best = max(
            (len(str(c.value)) for c in col if c.value is not None),
            default=10,
        )
        ws.column_dimensions[letter].width = min(best + 4, cap)


def _add_sheet(wb, name, columns, rows):
    ws = wb.create_sheet(name)
    ws.append(columns)
    _apply_header(ws, len(columns))
    for row in rows:
        ws.append([row.get(c) for c in columns])
    _fit_columns(ws)
    ws.freeze_panes = "A2"


def build_workbook(dataset):
    wb = Workbook()
    wb.remove(wb.active)

    _add_sheet(wb, "Customer", CUSTOMER_COLS, dataset["customers"])
    _add_sheet(wb, "Accounts", ACCOUNT_COLS, dataset["accounts"])
    _add_sheet(wb, "Transactions", TRANSACTION_COLS, dataset["transactions"])

    ws_dd = wb.create_sheet("Data_Dictionary")
    ws_dd.append(["Table Name", "Column Name", "Required Data Type", "Pandas Dtype"])
    _apply_header(ws_dd, 4)
    for row in DATA_DICT:
        ws_dd.append(list(row))
    _fit_columns(ws_dd, cap=40)
    ws_dd.freeze_panes = "A2"

    return wb


def workbook_to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
